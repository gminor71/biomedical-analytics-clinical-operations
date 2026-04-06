from __future__ import annotations

import argparse
import re
import yaml
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


def load_yaml(file_path: Path) -> dict:
    with open(file_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def slugify_project_name(project_name: str) -> str:
    slug = project_name.strip().lower()
    slug = re.sub(r"[^a-z0-9]+", "_", slug)
    slug = re.sub(r"_+", "_", slug).strip("_")
    return slug or "new_project"


def read_text(file_path: Path) -> str:
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()


def write_text(file_path: Path, content: str) -> None:
    file_path.parent.mkdir(parents=True, exist_ok=True)
    with open(file_path, "w", encoding="utf-8") as f:
        f.write(content)


def prefill_template_content(content: str, project_name: str) -> str:
    replacements = [
        ("- **Project Name:**  ", f"- **Project Name:** {project_name}"),
        ("- **Project Name:** ", f"- **Project Name:** {project_name}"),
        ("**Project Name:**  ", f"**Project Name:** {project_name}"),
        ("**Project Name:** ", f"**Project Name:** {project_name}"),
    ]

    updated = content
    for old, new in replacements:
        updated = updated.replace(old, new)

    return updated


def build_markdown_file_name_map() -> dict[str, str]:
    today_str = datetime.today().strftime("%Y-%m-%d")

    return {
        "project_charter": "project_charter.md",
        "raid_log": "raid_log_template.md",
        "decision_log": "decision_log.md",
        "status_report": f"status_report_{today_str}.md",
        "deliverables_tracker": "deliverables_tracker_template.md",
        "communication_plan": "communication_plan.md",
        "stakeholder_register": "stakeholder_register_template.md",
        "milestone_tracker": "milestone_tracker_template.md",
    }

def apply_title_style(cell) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_metadata_style(cell) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    cell.font = Font(italic=True)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def apply_header_style(cell) -> None:
    cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    cell.font = Font(bold=True, color="FFFFFF")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_column_widths(ws, widths: dict[str, float]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def add_list_validation(ws, cell_range: str, values: list[str]) -> None:
    quoted = ",".join(values)
    dv = DataValidation(type="list", formula1=f'"{quoted}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def add_standard_header(
    ws,
    title: str,
    project_name: str,
    source_name: str,
    merge_range_title: str,
    merge_range_meta: str,
) -> None:
    today_str = datetime.today().strftime("%Y-%m-%d")

    ws.merge_cells(merge_range_title)
    title_cell = ws["A1"]
    title_cell.value = f"{title} - {project_name}"
    apply_title_style(title_cell)

    ws.merge_cells(merge_range_meta)
    meta_cell = ws["A2"]
    meta_cell.value = f"Generated from PM01: {today_str} | Source: {source_name}"
    apply_metadata_style(meta_cell)


def finalize_sheet(ws, header_row: int, freeze_cell: str = "A5") -> None:
    ws.freeze_panes = freeze_cell
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[header_row].height = 22

def build_excel_file_name_map() -> dict[str, str]:
    return {
        "raid_log": "raid_log_working.xlsx",
        "deliverables_tracker": "deliverables_tracker_working.xlsx",
        "milestone_tracker": "milestone_tracker_working.xlsx",
        "stakeholder_register": "stakeholder_register_working.xlsx",
    }

def resolve_input_path(input_value: str, results_dir: Path) -> Path:
    candidate = Path(input_value)

    if candidate.is_absolute():
        return candidate

    if candidate.exists():
        return candidate.resolve()

    return (results_dir / input_value).resolve()


def style_header(rng) -> None:
    rng.format = {
        "fill": "#1F4E78",
        "font": {"bold": True, "color": "#FFFFFF"},
        "horizontal_alignment": "center",
        "vertical_alignment": "center",
        "wrap_text": True,
    }


def add_list_validation(ws, cell_range: str, values: list[str]) -> None:
    quoted = ",".join(values)
    dv = DataValidation(type="list", formula1=f'"{quoted}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)

def add_standard_header(
    ws,
    title: str,
    project_name: str,
    source_name: str,
    merge_range_title: str,
    merge_range_meta: str,
) -> None:
    today_str = datetime.today().strftime("%Y-%m-%d")

    ws.merge_cells(merge_range_title)
    title_cell = ws["A1"]
    title_cell.value = f"{title} - {project_name}"
    apply_title_style(title_cell)

    ws.merge_cells(merge_range_meta)
    meta_cell = ws["A2"]
    meta_cell.value = f"Generated from PM01: {today_str} | Source: {source_name}"
    apply_metadata_style(meta_cell)

def create_markdown_project_files(
    pm01_output_path: Path,
    project_name: str,
    required_outputs: set[str],
    recommended_outputs: set[str],
    template_refs: dict[str, str],
    required_dir: Path,
    recommended_dir: Path,
) -> None:
    file_name_map = build_markdown_file_name_map()

    for output_name, template_path_str in template_refs.items():
        template_path = (pm01_output_path.parent / template_path_str).resolve()

        if not template_path.exists():
            print(f"Template not found for {output_name}: {template_path}")
            continue

        if output_name in required_outputs:
            target_dir = required_dir
        elif output_name in recommended_outputs:
            target_dir = recommended_dir
        else:
            continue

        target_file_name = file_name_map.get(output_name, f"{output_name}.md")
        target_path = target_dir / target_file_name

        template_content = read_text(template_path)
        filled_content = prefill_template_content(
            content=template_content,
            project_name=project_name,
        )
        write_text(target_path, filled_content)

        print(f"Created markdown (template): {target_path}")


def build_raid_log_workbook(project_name: str, source_name: str, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "RAID Log"

    add_standard_header(
        ws=ws,
        title="RAID Log",
        project_name=project_name,
        source_name=source_name,
        merge_range_title="A1:N1",
        merge_range_meta="A2:N2",
    )

    headers = [
        "id", "type", "title", "description", "impact", "likelihood",
        "priority", "owner", "status", "created_date", "due_date",
        "source", "phase", "notes"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_header_style(cell)

    starter_rows = [
        ["RAID-001", "", "", "", "", "", "", "", "Open", "", "", "", "", ""],
        ["RAID-002", "", "", "", "", "", "", "", "Open", "", "", "", "", ""],
        ["RAID-003", "", "", "", "", "", "", "", "Open", "", "", "", "", ""],
    ]

    for row_idx, row_data in enumerate(starter_rows, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    add_list_validation(ws, "B5:B200", ["Risk", "Issue", "Action", "Decision"])
    add_list_validation(ws, "E5:E200", ["Low", "Medium", "High", "Critical"])
    add_list_validation(ws, "F5:F200", ["Low", "Medium", "High"])
    add_list_validation(ws, "G5:G200", ["Low", "Medium", "High", "Critical"])
    add_list_validation(ws, "I5:I200", ["Open", "In Progress", "Blocked", "Closed"])
    add_list_validation(ws, "M5:M200", ["Planning", "Execution", "Monitoring", "Closeout"])

    for row in range(5, 201):
        ws[f"J{row}"].number_format = "yyyy-mm-dd"
        ws[f"K{row}"].number_format = "yyyy-mm-dd"

    set_column_widths(ws, {
        "A": 12, "B": 14, "C": 24, "D": 36, "E": 12, "F": 12, "G": 12,
        "H": 18, "I": 14, "J": 14, "K": 14, "L": 18, "M": 14, "N": 28,
    })

    finalize_sheet(ws, header_row=4)
    wb.save(output_path)


def build_deliverables_tracker_workbook(project_name: str, source_name: str, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Deliverables"

    add_standard_header(
        ws=ws,
        title="Deliverables Tracker",
        project_name=project_name,
        source_name=source_name,
        merge_range_title="A1:H1",
        merge_range_meta="A2:H2",
    )

    headers = [
        "deliverable_id", "deliverable_name", "description", "owner",
        "planned_due_date", "current_forecast", "status", "notes"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_header_style(cell)

    starter_rows = [
        ["DEL-001", "", "", "", "", "", "Not Started", ""],
        ["DEL-002", "", "", "", "", "", "In Progress", ""],
        ["DEL-003", "", "", "", "", "", "Complete", ""],
    ]

    for row_idx, row_data in enumerate(starter_rows, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    add_list_validation(ws, "G5:G200", ["Not Started", "In Progress", "At Risk", "Blocked", "Complete"])

    for row in range(5, 201):
        ws[f"E{row}"].number_format = "yyyy-mm-dd"
        ws[f"F{row}"].number_format = "yyyy-mm-dd"

    set_column_widths(ws, {
        "A": 16, "B": 28, "C": 36, "D": 18, "E": 14, "F": 16, "G": 16, "H": 28,
    })

    finalize_sheet(ws, header_row=4)
    wb.save(output_path)


def build_milestone_tracker_workbook(project_name: str, source_name: str, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Milestones"

    add_standard_header(
        ws=ws,
        title="Milestone Tracker",
        project_name=project_name,
        source_name=source_name,
        merge_range_title="A1:H1",
        merge_range_meta="A2:H2",
    )

    headers = [
        "milestone_id", "milestone_name", "description", "owner",
        "planned_date", "current_forecast", "status", "impact_if_delayed"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_header_style(cell)

    starter_rows = [
        ["MS-001", "", "", "", "", "", "On Track", ""],
        ["MS-002", "", "", "", "", "", "At Risk", ""],
        ["MS-003", "", "", "", "", "", "Complete", ""],
    ]

    for row_idx, row_data in enumerate(starter_rows, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    add_list_validation(ws, "G5:G200", ["On Track", "At Risk", "Delayed", "Complete"])

    for row in range(5, 201):
        ws[f"E{row}"].number_format = "yyyy-mm-dd"
        ws[f"F{row}"].number_format = "yyyy-mm-dd"

    set_column_widths(ws, {
        "A": 14, "B": 28, "C": 34, "D": 18, "E": 14, "F": 16, "G": 14, "H": 28,
    })

    finalize_sheet(ws, header_row=4)
    wb.save(output_path)


def build_stakeholder_register_workbook(project_name: str, source_name: str, output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stakeholders"

    add_standard_header(
        ws=ws,
        title="Stakeholder Register",
        project_name=project_name,
        source_name=source_name,
        merge_range_title="A1:G1",
        merge_range_meta="A2:G2",
    )

    headers = [
        "stakeholder_id", "name_or_group", "role", "function",
        "level_of_influence", "level_of_interest", "engagement_strategy"
    ]

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        apply_header_style(cell)

    starter_rows = [
        ["STK-001", "", "", "", "High", "High", ""],
        ["STK-002", "", "", "", "Medium", "High", ""],
        ["STK-003", "", "", "", "Low", "Medium", ""],
    ]

    for row_idx, row_data in enumerate(starter_rows, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    add_list_validation(ws, "E5:E200", ["Low", "Medium", "High"])
    add_list_validation(ws, "F5:F200", ["Low", "Medium", "High"])

    set_column_widths(ws, {
        "A": 16, "B": 28, "C": 20, "D": 20, "E": 18, "F": 18, "G": 30,
    })

    finalize_sheet(ws, header_row=4)
    wb.save(output_path)

def create_excel_project_files(
    project_name: str,
    source_name: str,
    required_outputs: set[str],
    recommended_outputs: set[str],
    required_dir: Path,
    recommended_dir: Path,
) -> None:
    excel_file_map = build_excel_file_name_map()

    builders = {
        "raid_log": build_raid_log_workbook,
        "deliverables_tracker": build_deliverables_tracker_workbook,
        "milestone_tracker": build_milestone_tracker_workbook,
        "stakeholder_register": build_stakeholder_register_workbook,
    }

    for output_name, file_name in excel_file_map.items():
        if output_name in required_outputs:
            target_dir = required_dir
        elif output_name in recommended_outputs:
            target_dir = recommended_dir
        else:
            continue

        target_path = target_dir / file_name
        builders[output_name](project_name, source_name, target_path)
        print(f"Created excel (working): {target_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate project files from a PM01 recommender output YAML."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="PM01 output YAML file name or full path.",
    )
    parser.add_argument(
        "--output-root",
        default="generated_projects",
        help="Output folder name or full path for generated project directories.",
    )
    parser.add_argument(
        "--format",
        choices=["markdown", "excel", "both"],
        default="both",
        help="Which output format to generate.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    base_dir = Path(__file__).resolve().parent
    results_dir = base_dir / "results"

    input_file = resolve_input_path(args.input, results_dir)

    if not input_file.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}")

    output_root_candidate = Path(args.output_root)
    if output_root_candidate.is_absolute():
        output_root = output_root_candidate
    else:
        output_root = (base_dir / output_root_candidate).resolve()

    data = load_yaml(input_file)

    project_name = data.get("project_name", "New Project")
    project_slug = slugify_project_name(project_name)
    source_name = input_file.name

    project_dir = output_root / project_slug
    required_dir = project_dir / "required"
    recommended_dir = project_dir / "recommended"

    required_dir.mkdir(parents=True, exist_ok=True)
    recommended_dir.mkdir(parents=True, exist_ok=True)

    required_outputs = set(data.get("required_outputs", []))
    recommended_outputs = set(data.get("recommended_outputs", []))
    template_refs = data.get("template_references", {})

    if args.format in {"markdown", "both"}:
        create_markdown_project_files(
            pm01_output_path=input_file,
            project_name=project_name,
            required_outputs=required_outputs,
            recommended_outputs=recommended_outputs,
            template_refs=template_refs,
            required_dir=required_dir,
            recommended_dir=recommended_dir,
        )

    if args.format in {"excel", "both"}:
        create_excel_project_files(
            project_name=project_name,
            source_name=source_name,
            required_outputs=required_outputs,
            recommended_outputs=recommended_outputs,
            required_dir=required_dir,
            recommended_dir=recommended_dir,
        )

    print(f"\nProject files created in: {project_dir}")


if __name__ == "__main__":
    main()