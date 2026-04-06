from __future__ import annotations

import argparse
import re
import yaml
from pathlib import Path

from openpyxl import load_workbook

from datetime import datetime

def load_yaml(file_path: Path) -> dict:
    with open(file_path, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def resolve_input_path(input_value: str, base_dir: Path) -> Path:
    candidate = Path(input_value)

    if candidate.is_absolute():
        return candidate

    if candidate.exists():
        return candidate.resolve()

    return (base_dir / input_value).resolve()


def get_next_raid_id(ws) -> str:
    pattern = re.compile(r"RAID-(\d+)")
    max_id = 0

    for row in range(5, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if isinstance(value, str):
            match = pattern.fullmatch(value.strip())
            if match:
                max_id = max(max_id, int(match.group(1)))

    return f"RAID-{max_id + 1:03d}"


def find_next_empty_row(ws) -> int:
    row = 5
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1
    return row


def append_triaged_item_to_raid_log(pm02_result_path: Path, raid_log_path: Path) -> None:
    data = load_yaml(pm02_result_path)

    wb = load_workbook(raid_log_path)
    ws = wb["RAID Log"]

    next_row = find_next_empty_row(ws)
    new_id = get_next_raid_id(ws)

    ws.cell(row=next_row, column=1, value=new_id)
    ws.cell(row=next_row, column=2, value=data.get("classification", ""))
    ws.cell(row=next_row, column=3, value=data.get("item_title", ""))
    ws.cell(row=next_row, column=4, value=data.get("triage_note", ""))
    ws.cell(row=next_row, column=5, value=data.get("priority", ""))
    ws.cell(row=next_row, column=6, value="")
    ws.cell(row=next_row, column=7, value=data.get("priority", ""))
    ws.cell(row=next_row, column=8, value=data.get("owner_type", ""))
    ws.cell(row=next_row, column=9, value=data.get("suggested_status", "Open"))
    ws.cell(row=next_row, column=10, value=datetime.today().strftime("%Y-%m-%d"))
    ws.cell(row=next_row, column=11, value="")
    ws.cell(row=next_row, column=12, value=pm02_result_path.name)
    ws.cell(row=next_row, column=13, value=data.get("suggested_phase", "Execution"))
    ws.cell(row=next_row, column=14, value="; ".join(data.get("recommended_actions", [])))

    wb.save(raid_log_path)

    print(f"Appended {new_id} to {raid_log_path}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Append a PM02 triaged item to a RAID log workbook."
    )
    parser.add_argument(
        "--input",
        required=True,
        help="PM02 result YAML file name or full path.",
    )
    parser.add_argument(
        "--raid-log",
        required=True,
        help="Path to raid_log_working.xlsx",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()

    base_dir = Path(__file__).resolve().parent
    results_dir = base_dir / "results"

    pm02_result_path = resolve_input_path(args.input, results_dir)
    raid_log_path = resolve_input_path(args.raid_log, base_dir)

    if not pm02_result_path.exists():
        raise FileNotFoundError(f"PM02 result file not found: {pm02_result_path}")

    if not raid_log_path.exists():
        raise FileNotFoundError(f"RAID log workbook not found: {raid_log_path}")

    append_triaged_item_to_raid_log(pm02_result_path, raid_log_path)


if __name__ == "__main__":
    main()